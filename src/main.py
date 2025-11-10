"""
Основной модуль проекта GENERATE_RANDOM_IFT_USER_12M_4Q_BASE.

Точка входа в приложение.
Весь код проекта находится в этом файле, разделенный на модули.
Использует библиотеки из стандартной установки Anaconda (pandas, openpyxl).
"""

# ============================================================================
# ПАРАМЕТРЫ КОНФИГУРАЦИИ
# ============================================================================

# Настройки приложения
APP_NAME = "GENERATE_RANDOM_IFT_USER_12M_4Q_BASE"  # Имя приложения
DEBUG = True  # Режим отладки

# Параметры логирования
LOG_DIR = "log"  # Директория для хранения логов
LOG_LEVEL = "DEBUG"  # Уровень логирования: INFO или DEBUG

# Общие параметры выходных файлов
OUTPUT_DIR = "OUT"  # Директория для выходных Excel файлов
OUTPUT_FILE_BASE = "result_base"  # Базовое имя выходного Excel файла

# ============================================================================
# КОНФИГУРАЦИЯ ЗАГРУЗЧИКОВ ДАННЫХ
# ============================================================================
# Структура конфигурации для каждого загрузчика данных
# Каждый загрузчик соответствует отдельному листу в Excel файле

LOADER_CONFIG = {
    'ORG': {
        # Входной файл
        'input_file': "IN/SVD_KB_DM_GAMIFICATION_ORG_UNIT_V20 - 2025.08.28.csv",
        
        # Настройки выходного листа
        'sheet_name': 'ORG',  # Имя листа в Excel
        'max_column_width': 100,  # Максимальная ширина колонки
        
        # Фильтры для исключения строк
        'filters': {
            'tb_code_exclude': {'99', '100', '101', '102'},  # Коды ТБ для исключения
            'gosb_code_exclude': {'0', '9038', '9040'}  # Коды ГОСБ для исключения
        },
        
        # Маппинг колонок: исходное_имя -> результирующее_имя
        'column_mapping': {
            'TB_CODE': 'Код ТБ',
            'TB_FULL_NAME': 'Полное ТБ',
            'TB_SHORT_NAME': 'Короткое ТБ',
            'GOSB_CODE': 'Код ГОСБ',
            'GOSB_NAME': 'Полное ГОСБ',
            'GOSB_SHORT_NAME': 'Короткое ГОСБ',
            'ORG_UNIT_CODE': 'Код подразделения'
        }
    },
    
    'USERS': {
        # Настройки выходного листа
        'sheet_name': 'USERS',  # Имя листа в Excel
        'max_column_width': 100,  # Максимальная ширина колонки
        
        # Параметры генерации пользователей по блокам
        'business_blocks': {
            'KMKKSB': {
                'name': 'Клиентские менеджеры',
                'count': 1600,
                'gender_distribution': 0.7  # 50% мужчин, 50% женщин
            },
            'MNS': {
                'name': 'Менеджер нефинансовых сервисов',
                'count': 160,
                'gender_distribution': 0.2  # 50% мужчин, 50% женщин
            }
        },
        
        # Параметры табельных номеров
        'tab_number': {
            'min_digits': 4,  # Минимальное количество значащих цифр
            'max_digits': 7,  # Максимальное количество значащих цифр
            'total_length': 8,  # Общая длина с лидирующими нулями
            'start_from': 1000  # Начальное значение (больше 999)
        },
        
        # Параметры специальных пользователей "Серая зона"
        'gray_zone': {
            'tab_numbers': ["gray_zone", "00000000", "-00000001"],  # Варианты табельных номеров
            'fio_options': ["Серая зона", "-"]  # Варианты ФИО
        },
        
        # Фиксированное распределение пользователей по подразделениям
        # Формат: {'код_подразделения': {'KMKKSB': количество, 'MNS': количество}}
        # Пример: {'12345': {'KMKKSB': 5, 'MNS': 1}, '67890': {'KMKKSB': 50}}
        # Если подразделение не указано, пользователи распределяются случайно
        'fixed_distribution': {
            # Примеры (закомментированы):
            # '12345': {'KMKKSB': 5, 'MNS': 1},  # В подразделение 12345: 5 КМ и 1 МНС
            # '67890': {'KMKKSB': 50},  # В подразделение 67890: 50 КМ
            '10214308': {'KMKKSB': 1, 'MNS': 1},
            '10354600': {'KMKKSB': 2, 'MNS': 1}
        },
        
        # Данные для генерации мужских ФИО
        'male_data': {
            'first_names': [
                'Александр', 'Дмитрий', 'Максим', 'Сергей', 'Андрей', 'Алексей', 'Артем', 'Илья',
                'Кирилл', 'Михаил', 'Никита', 'Матвей', 'Роман', 'Егор', 'Арсений', 'Иван',
                'Денис', 'Евгений', 'Данил', 'Тимур', 'Владислав', 'Игорь', 'Владимир', 'Павел',
                'Руслан', 'Марк', 'Лев', 'Андрей', 'Ярослав', 'Федор', 'Глеб', 'Николай',
                'Степан', 'Василий', 'Юрий', 'Борис', 'Олег', 'Константин', 'Виктор', 'Петр'
            ],
            'surnames': [
                'Иванов', 'Петров', 'Смирнов', 'Козлов', 'Попов', 'Соколов', 'Лебедев', 'Новиков',
                'Морозов', 'Волков', 'Соловьев', 'Васильев', 'Зайцев', 'Павлов', 'Семенов',
                'Голубев', 'Виноградов', 'Богданов', 'Воробьев', 'Федоров', 'Михайлов', 'Белов', 'Тарасов',
                'Беляев', 'Комаров', 'Орлов', 'Киселев', 'Макаров', 'Андреев', 'Ковалев', 'Ильин',
                'Гусев', 'Титов', 'Кузьмин', 'Кудрявцев', 'Баранов', 'Куликов', 'Алексеев', 'Степанов',
                'Яковлев', 'Сорокин', 'Сергеев', 'Романов', 'Захаров', 'Борисов', 'Королев', 'Герасимов',
                'Пономарев', 'Григорьев', 'Лазарев', 'Медведев', 'Ершов', 'Никитин', 'Соболев', 'Рябов',
                'Поляков', 'Цветков', 'Данилов', 'Жуков', 'Фролов', 'Журавлев', 'Николаев', 'Крылов',
                'Максимов', 'Сидоров', 'Осипов', 'Белоусов', 'Федотов', 'Дорофеев', 'Егоров', 'Матвеев',
                'Бобров', 'Дмитриев', 'Калинин', 'Анисимов', 'Петухов', 'Антонов', 'Тимофеев', 'Никифоров',
                'Веселов', 'Филиппов', 'Марков', 'Большаков', 'Суханов', 'Миронов', 'Ширяев', 'Александров',
                'Коновалов', 'Шестаков', 'Казаков', 'Ефимов', 'Денисов', 'Громов', 'Фомин', 'Давыдов',
                'Мельников', 'Щербаков', 'Блинов', 'Колесников', 'Карпов', 'Афанасьев', 'Власов', 'Маслов',
                'Исаков', 'Тихонов', 'Аксенов', 'Гаврилов', 'Родионов', 'Котов', 'Горбунов', 'Кудряшов',
                'Быков', 'Зуев', 'Третьяков', 'Савельев', 'Панов', 'Рыбаков', 'Суворов', 'Абрамов', 'Воронов',
                'Мухин', 'Архипов', 'Трофимов', 'Мартынов', 'Емельянов', 'Горшков', 'Чернов', 'Овчинников',
                'Селезнев', 'Панфилов', 'Копылов', 'Михеев', 'Галкин', 'Назаров', 'Лобанов', 'Лукин',
                'Беляков', 'Потапов', 'Некрасов', 'Хохлов', 'Жданов', 'Наумов', 'Шилов', 'Воронцов',
                'Ермаков', 'Дроздов', 'Игнатьев', 'Савин', 'Логинов', 'Сафонов', 'Капустин', 'Кириллов',
                'Моисеев', 'Елисеев', 'Кошелев', 'Костин', 'Горбачев', 'Орехов', 'Ефремов', 'Исаев',
                'Евдокимов', 'Калашников', 'Кабанов', 'Носков', 'Юдин', 'Кулагин', 'Лапин', 'Прохоров',
                'Нестеров', 'Харитонов', 'Агафонов', 'Муравьев', 'Ларионов', 'Федосеев', 'Зимин', 'Пахомов',
                'Шубин', 'Игнатов', 'Филатов', 'Крюков', 'Рогов', 'Кулаков', 'Терентьев', 'Молчанов',
                'Владимиров', 'Артемьев', 'Гурьев', 'Зиновьев', 'Гришин', 'Кононов', 'Дементьев', 'Ситников',
                'Симонов', 'Мишин', 'Фадеев', 'Комиссаров', 'Мамонтов', 'Носов', 'Гуляев', 'Шаров',
                'Устинов', 'Вишняков', 'Евсеев', 'Лаврентьев', 'Брагин', 'Константинов', 'Корнилов', 'Авдеев',
                'Зыков', 'Бирюков', 'Шарапов', 'Никонов', 'Щукин', 'Дьячков', 'Одинцов', 'Сазонов',
                'Якушев', 'Красильников', 'Гордеев', 'Самойлов', 'Князев', 'Беспалов', 'Уваров', 'Шашков',
                'Бобылев', 'Доронин', 'Белозеров', 'Рожков', 'Самсонов', 'Мясников', 'Лихачев', 'Буров',
                'Сысоев', 'Фомичев', 'Русаков', 'Стрелков', 'Гущин', 'Тетерин', 'Колобов', 'Субботин',
                'Фокин', 'Блохин', 'Селиверстов', 'Пестов', 'Кондратьев', 'Силин', 'Меркушев', 'Лыткин',
                'Туров'
            ],
            'patronymics': [
                'Александрович', 'Дмитриевич', 'Максимович', 'Сергеевич', 'Андреевич',
                'Алексеевич', 'Артемович', 'Ильич', 'Кириллович', 'Михайлович',
                'Никитич', 'Матвеевич', 'Романович', 'Егорович', 'Арсеньевич',
                'Иванович', 'Денисович', 'Евгеньевич', 'Данилович', 'Тимурович',
                'Владиславович', 'Игоревич', 'Владимирович', 'Павлович', 'Русланович'
            ]
        },
        
        # Данные для генерации женских ФИО
        'female_data': {
            'first_names': [
                'Анна', 'Мария', 'Елена', 'Наталья', 'Ольга', 'Татьяна', 'Ирина', 'Екатерина',
                'Светлана', 'Марина', 'Надежда', 'Евгения', 'Юлия', 'Анастасия', 'Оксана', 'Анжела',
                'Валентина', 'Вера', 'Галина', 'Людмила', 'Лариса', 'Алина', 'Дарья', 'София',
                'Виктория', 'Полина', 'Алиса', 'Ксения', 'Валерия', 'Диана', 'Арина', 'Милана',
                'Варвара', 'Ульяна', 'Маргарита', 'Ангелина', 'Елизавета', 'Василиса', 'Амелия', 'Вероника'
            ],
            'surnames': [
                'Иванова', 'Петрова', 'Смирнова', 'Козлова', 'Попова', 'Соколова', 'Лебедева', 'Новикова',
                'Морозова', 'Волкова', 'Соловьева', 'Васильева', 'Зайцева', 'Павлова', 'Семенова',
                'Голубева', 'Виноградова', 'Богданова', 'Воробьева', 'Федорова', 'Михайлова', 'Белова', 'Тарасова',
                'Беляева', 'Комарова', 'Орлова', 'Киселева', 'Макарова', 'Андреева', 'Ковалева', 'Ильина',
                'Гусева', 'Титова', 'Кузьмина', 'Кудрявцева', 'Баранова', 'Куликова', 'Алексеева', 'Степанова',
                'Яковлева', 'Сорокина', 'Сергеева', 'Романова', 'Захарова', 'Борисова', 'Королева', 'Герасимова',
                'Пономарева', 'Григорьева', 'Лазарева', 'Медведева', 'Ершова', 'Никитина', 'Соболева', 'Рябова',
                'Полякова', 'Цветкова', 'Данилова', 'Жукова', 'Фролова', 'Журавлева', 'Николаева', 'Крылова',
                'Максимова', 'Сидорова', 'Осипова', 'Белоусова', 'Федотова', 'Дорофеева', 'Егорова', 'Матвеева',
                'Боброва', 'Дмитриева', 'Калинина', 'Анисимова', 'Петухова', 'Антонова', 'Тимофеева', 'Никифорова',
                'Веселова', 'Филиппова', 'Маркова', 'Большакова', 'Суханова', 'Миронова', 'Ширяева', 'Александрова',
                'Коновалова', 'Шестакова', 'Казакова', 'Ефимова', 'Денисова', 'Громова', 'Фомина', 'Давыдова',
                'Мельникова', 'Щербакова', 'Блинова', 'Колесникова', 'Карпова', 'Афанасьева', 'Власова', 'Маслова',
                'Исакова', 'Тихонова', 'Аксенова', 'Гаврилова', 'Родионова', 'Котова', 'Горбунова', 'Кудряшова',
                'Быкова', 'Зуева', 'Третьякова', 'Савельева', 'Панова', 'Рыбакова', 'Суворова', 'Абрамова', 'Воронова',
                'Мухина', 'Архипова', 'Трофимова', 'Мартынова', 'Емельянова', 'Горшкова', 'Чернова', 'Овчинникова',
                'Селезнева', 'Панфилова', 'Копылова', 'Михеева', 'Галкина', 'Назарова', 'Лобанова', 'Лукина',
                'Белякова', 'Потапова', 'Некрасова', 'Хохлова', 'Жданова', 'Наумова', 'Шилова', 'Воронцова',
                'Ермакова', 'Дроздова', 'Игнатьева', 'Савина', 'Логинова', 'Сафонова', 'Капустина', 'Кириллова',
                'Моисеева', 'Елисеева', 'Кошелева', 'Костина', 'Горбачева', 'Орехова', 'Ефремова', 'Исаева',
                'Евдокимова', 'Калашникова', 'Кабанова', 'Носкова', 'Юдина', 'Кулагина', 'Лапина', 'Прохорова',
                'Нестерова', 'Харитонова', 'Агафонова', 'Муравьева', 'Ларионова', 'Федосеева', 'Зимина', 'Пахомова',
                'Шубина', 'Игнатова', 'Филатова', 'Крюкова', 'Рогова', 'Кулакова', 'Терентьева', 'Молчанова',
                'Владимирова', 'Артемьева', 'Гурьева', 'Зиновьева', 'Гришина', 'Кононова', 'Дементьева', 'Ситникова',
                'Симонова', 'Мишина', 'Фадеева', 'Комиссарова', 'Мамонтова', 'Носова', 'Гуляева', 'Шарова',
                'Устинова', 'Вишнякова', 'Евсеева', 'Лаврентьева', 'Брагина', 'Константинова', 'Корнилова', 'Авдеева',
                'Зыкова', 'Бирюкова', 'Шарапова', 'Никонова', 'Щукина', 'Дьячкова', 'Одинцова', 'Сазонова',
                'Якушева', 'Красильникова', 'Гордеева', 'Самойлова', 'Князева', 'Беспалова', 'Уварова', 'Шашкова',
                'Бобылева', 'Доронина', 'Белозерова', 'Рожкова', 'Самсонова', 'Мясникова', 'Лихачева', 'Бурова',
                'Сысоева', 'Фомичева', 'Русакова', 'Стрелкова', 'Гущина', 'Тетерина', 'Колобова', 'Субботина',
                'Фокина', 'Блохина', 'Селиверстова', 'Пестова', 'Кондратьева', 'Силина', 'Меркушева', 'Лыткина',
                'Турова'
            ],
            'patronymics': [
                'Александровна', 'Дмитриевна', 'Максимовна', 'Сергеевна', 'Андреевна',
                'Алексеевна', 'Артемовна', 'Ильинична', 'Кирилловна', 'Михайловна',
                'Никитична', 'Матвеевна', 'Романовна', 'Егоровна', 'Арсеньевна',
                'Ивановна', 'Денисовна', 'Евгеньевна', 'Даниловна', 'Тимуровна',
                'Владиславовна', 'Игоревна', 'Владимировна', 'Павловна', 'Руслановна'
            ]
        }
    }
    
    # Здесь в будущем можно добавить конфигурации для других листов:
    # 'METRICS': {
    #     'input_file': "IN/metrics.csv",
    #     'sheet_name': 'METRICS',
    #     'max_column_width': 100,
    #     'filters': {...},
    #     'column_mapping': {...}
    # }
}


# ============================================================================
# МОДУЛЬ ЛОГИРОВАНИЯ
# ============================================================================

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List


class ProjectLogger:
    """
    Класс для управления логированием проекта.
    
    Создает логи с форматом: Уровень_(тема)_годмесяцдень_часминута.log
    DEBUG логи имеют формат: дата время - [уровень] - сообщение [class: <имя класса> | def: <имя функции>]
    """
    
    def __init__(self, log_dir: str = "log", log_level: str = "DEBUG") -> None:
        """
        Инициализация логгера.
        
        Args:
            log_dir: Директория для хранения логов
            log_level: Уровень логирования (INFO или DEBUG)
        """
        self.log_dir = Path(log_dir)
        self.log_dir.mkdir(exist_ok=True)
        self.log_level = getattr(logging, log_level.upper(), logging.DEBUG)
        
        # Настройка форматирования для DEBUG
        debug_formatter = logging.Formatter(
            '%(asctime)s - [%(levelname)s] - %(message)s [class: %(name)s | def: %(funcName)s]',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        
        # Настройка форматирования для INFO
        info_formatter = logging.Formatter(
            '%(asctime)s - [%(levelname)s] - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        
        # Создание логгера
        self.logger = logging.getLogger('project_logger')
        self.logger.setLevel(self.log_level)
        
        # Очистка существующих обработчиков
        self.logger.handlers.clear()
        
        # Создание файловых обработчиков
        self._setup_file_handlers(debug_formatter, info_formatter)
        
        # Создание консольного обработчика для INFO
        self._setup_console_handler(info_formatter)
    
    def _setup_file_handlers(self, debug_formatter: logging.Formatter, info_formatter: logging.Formatter) -> None:
        """
        Настройка файловых обработчиков для логирования.
        
        Args:
            debug_formatter: Форматтер для DEBUG логов
            info_formatter: Форматтер для INFO логов
        """
        now = datetime.now()
        timestamp = now.strftime("%Y%m%d_%H%M")
        
        # Обработчик для DEBUG
        debug_file = self.log_dir / f"DEBUG_project_{timestamp}.log"
        debug_handler = logging.FileHandler(debug_file, encoding='utf-8')
        debug_handler.setLevel(logging.DEBUG)
        debug_handler.setFormatter(debug_formatter)
        self.logger.addHandler(debug_handler)
        
        # Обработчик для INFO
        info_file = self.log_dir / f"INFO_project_{timestamp}.log"
        info_handler = logging.FileHandler(info_file, encoding='utf-8')
        info_handler.setLevel(logging.INFO)
        info_handler.setFormatter(info_formatter)
        self.logger.addHandler(info_handler)
    
    def _setup_console_handler(self, info_formatter: logging.Formatter) -> None:
        """
        Настройка консольного обработчика для вывода сообщений уровня INFO.
        
        Args:
            info_formatter: Форматтер для INFO логов
        """
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(info_formatter)
        self.logger.addHandler(console_handler)
    
    def get_logger(self) -> logging.Logger:
        """
        Получение настроенного логгера.
        
        Returns:
            Настроенный объект логгера
        """
        return self.logger


def get_logger(log_dir: Optional[str] = None, log_level: Optional[str] = None) -> logging.Logger:
    """
    Функция для получения глобального логгера проекта.
    
    Args:
        log_dir: Директория для логов (по умолчанию из параметров конфигурации)
        log_level: Уровень логирования (по умолчанию из параметров конфигурации)
    
    Returns:
        Настроенный объект логгера
    """
    if log_dir is None:
        log_dir = LOG_DIR
    if log_level is None:
        log_level = LOG_LEVEL
    
    logger_instance = ProjectLogger(log_dir=log_dir, log_level=log_level)
    return logger_instance.get_logger()


# ============================================================================
# МОДУЛЬ ЗАГРУЗКИ ОРГАНИЗАЦИОННЫХ ЕДИНИЦ
# ============================================================================

import pandas as pd
import random
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class OrgUnitsLoader:
    """
    Класс для загрузки и обработки организационных единиц.
    
    Загружает данные из CSV, применяет фильтры и сохраняет в Excel.
    Использует библиотеки из стандартной установки Anaconda (pandas, openpyxl).
    """
    
    def __init__(
        self,
        config: Dict,
        output_file_base: str,
        output_dir: str = "OUT",
        logger: Optional[logging.Logger] = None
    ) -> None:
        """
        Инициализация загрузчика организационных единиц.
        
        Args:
            config: Словарь конфигурации из LOADER_CONFIG для данного загрузчика
            output_file_base: Базовое имя выходного Excel файла (без расширения)
            output_dir: Директория для выходных файлов
            logger: Логгер для записи событий
        """
        self.config = config
        self.input_file = Path(config['input_file'])
        self.output_file_base = output_file_base
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)  # Создаем папку, если её нет
        self.logger = logger or logging.getLogger(__name__)
        
        # Маппинг колонок: исходное_имя -> результирующее_имя
        self.column_mapping = config['column_mapping']
        
        # Фильтры для исключения строк
        filters = config.get('filters', {})
        self.tb_code_exclude = filters.get('tb_code_exclude', set())
        self.gosb_code_exclude = filters.get('gosb_code_exclude', set())
        
        # Настройки Excel
        self.sheet_name = config['sheet_name']
        self.max_column_width = config.get('max_column_width', 100)
    
    def load_csv(self) -> pd.DataFrame:
        """
        Загрузка данных из CSV файла.
        
        Returns:
            DataFrame с загруженными данными
            
        Raises:
            FileNotFoundError: Если файл не найден
            ValueError: Если файл пустой или неверный формат
        """
        if not self.input_file.exists():
            error_msg = f"Файл не найден: {self.input_file}"
            self.logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        self.logger.info(f"Начало загрузки CSV файла: {self.input_file}")
        self.logger.debug(f"Загрузка данных из файла [class: OrgUnitsLoader | def: load_csv]")
        
        try:
            # Загрузка CSV с разделителем ';'
            df = pd.read_csv(
                self.input_file,
                sep=';',
                encoding='utf-8',
                dtype=str  # Загружаем все как строки для корректной фильтрации
            )
            
            self.logger.info(f"Загружено строк: {len(df)}")
            self.logger.debug(f"Загружено {len(df)} строк, колонок: {len(df.columns)} [class: OrgUnitsLoader | def: load_csv]")
            
            if df.empty:
                raise ValueError("CSV файл пуст")
            
            return df
            
        except Exception as e:
            error_msg = f"Ошибка при загрузке CSV: {str(e)}"
            self.logger.error(error_msg)
            raise
    
    def filter_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Фильтрация данных по заданным критериям.
        
        Исключает строки где:
        - TB_CODE в списке исключений (99, 100, 101, 102)
        - GOSB_CODE в списке исключений (0, 9038, 9040)
        
        Args:
            df: Исходный DataFrame
            
        Returns:
            Отфильтрованный DataFrame
        """
        initial_count = len(df)
        self.logger.debug(f"Начало фильтрации данных. Исходное количество строк: {initial_count} [class: OrgUnitsLoader | def: filter_data]")
        
        # Проверяем наличие необходимых колонок
        required_columns = ['TB_CODE', 'GOSB_CODE']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            error_msg = f"Отсутствуют необходимые колонки: {missing_columns}"
            self.logger.error(error_msg)
            raise ValueError(error_msg)
        
        # Фильтрация по TB_CODE
        df_filtered = df.copy()
        excluded_tb = 0
        if 'TB_CODE' in df.columns:
            count_before = len(df_filtered)
            df_filtered = df_filtered[~df_filtered['TB_CODE'].isin(self.tb_code_exclude)]
            excluded_tb = count_before - len(df_filtered)
            if excluded_tb > 0:
                self.logger.debug(f"Исключено строк по TB_CODE: {excluded_tb} [class: OrgUnitsLoader | def: filter_data]")
        
        # Фильтрация по GOSB_CODE
        excluded_gosb = 0
        if 'GOSB_CODE' in df_filtered.columns:
            count_before = len(df_filtered)
            df_filtered = df_filtered[~df_filtered['GOSB_CODE'].isin(self.gosb_code_exclude)]
            excluded_gosb = count_before - len(df_filtered)
            if excluded_gosb > 0:
                self.logger.debug(f"Исключено строк по GOSB_CODE: {excluded_gosb} [class: OrgUnitsLoader | def: filter_data]")
        
        final_count = len(df_filtered)
        excluded_total = initial_count - final_count
        
        self.logger.info(f"Отфильтровано строк: {excluded_total}, осталось: {final_count}")
        self.logger.debug(f"Фильтрация завершена. Исключено: {excluded_total}, осталось: {final_count} [class: OrgUnitsLoader | def: filter_data]")
        
        return df_filtered
    
    def select_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Выбор и переименование необходимых колонок.
        
        Args:
            df: Исходный DataFrame
            
        Returns:
            DataFrame с выбранными и переименованными колонками
        """
        self.logger.debug(f"Выбор колонок для экспорта [class: OrgUnitsLoader | def: select_columns]")
        
        # Проверяем наличие всех необходимых колонок
        missing_columns = [col for col in self.column_mapping.keys() if col not in df.columns]
        if missing_columns:
            error_msg = f"Отсутствуют необходимые колонки в исходных данных: {missing_columns}"
            self.logger.error(error_msg)
            raise ValueError(error_msg)
        
        # Выбираем только нужные колонки
        selected_df = df[list(self.column_mapping.keys())].copy()
        
        # Переименовываем колонки
        selected_df = selected_df.rename(columns=self.column_mapping)
        
        self.logger.debug(f"Выбрано и переименовано колонок: {len(selected_df.columns)} [class: OrgUnitsLoader | def: select_columns]")
        
        return selected_df
    
    def save_to_excel(self, df: pd.DataFrame) -> str:
        """
        Сохранение данных в Excel файл с настройками форматирования.
        
        Настройки:
        - Первая строка закреплена
        - Автофильтр включен
        - Ширина колонок по содержимому (максимум max_column_width)
        
        Args:
            df: DataFrame для сохранения
            
        Returns:
            Путь к созданному файлу
        """
        # Формируем имя файла с таймштампом
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_file = f"{self.output_file_base}_{timestamp}.xlsx"
        output_path = self.output_dir / output_file
        
        self.logger.info(f"Сохранение данных в Excel: {output_file}")
        self.logger.debug(f"Сохранение {len(df)} строк в файл {output_file} [class: OrgUnitsLoader | def: save_to_excel]")
        
        # Сохраняем в Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=self.sheet_name, index=False)
            
            # Получаем объект листа для форматирования
            worksheet = writer.sheets[self.sheet_name]
            
            # Закрепляем первую строку
            worksheet.freeze_panes = 'A2'
            
            # Включаем автофильтр
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # Настраиваем ширину колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                # Находим максимальную длину содержимого в колонке
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                # Устанавливаем ширину (минимум 10, максимум max_column_width)
                adjusted_width = min(max(max_length + 2, 10), self.max_column_width)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Делаем первую строку жирной (заголовки)
            header_font = Font(bold=True)
            for cell in worksheet[1]:
                cell.font = header_font
        
        self.logger.info(f"Файл успешно создан: {output_path.absolute()}")
        self.logger.debug(f"Excel файл создан: {output_path.absolute()}, строк: {len(df)} [class: OrgUnitsLoader | def: save_to_excel]")
        
        return str(output_path.absolute())
    
    def process(self) -> str:
        """
        Полный цикл обработки: загрузка, фильтрация, выбор колонок, сохранение.
        
        Returns:
            Путь к созданному Excel файлу
        """
        self.logger.info("Начало обработки организационных единиц")
        self.logger.debug("Запуск полного цикла обработки данных [class: OrgUnitsLoader | def: process]")
        
        # Загрузка
        df = self.load_csv()
        
        # Фильтрация
        df_filtered = self.filter_data(df)
        
        # Выбор колонок
        df_final = self.select_columns(df_filtered)
        
        # Сохранение
        output_file = self.save_to_excel(df_final)
        
        self.logger.info(f"Обработка завершена успешно. Файл: {output_file}")
        self.logger.debug(f"Обработка завершена. Создан файл: {output_file} [class: OrgUnitsLoader | def: process]")
        
        return output_file


def load_org_units(
    config: Dict,
    output_file_base: str = "result_base",
    output_dir: str = "OUT",
    logger: Optional[logging.Logger] = None
) -> str:
    """
    Функция для загрузки и обработки организационных единиц.
    
    Args:
        config: Словарь конфигурации из LOADER_CONFIG для загрузчика
        output_file_base: Базовое имя выходного Excel файла (по умолчанию "result_base")
        output_dir: Директория для выходных файлов (по умолчанию "OUT")
        logger: Логгер для записи событий (опционально)
    
    Returns:
        Путь к созданному Excel файлу
        
    Example:
        >>> output = load_org_units(
        ...     config=LOADER_CONFIG['ORG'],
        ...     output_file_base="result_base",
        ...     output_dir="OUT"
        ... )
        >>> print(f"Создан файл: {output}")
    """
    loader = OrgUnitsLoader(
        config=config,
        output_file_base=output_file_base,
        output_dir=output_dir,
        logger=logger
    )
    return loader.process()


# ============================================================================
# МОДУЛЬ ГЕНЕРАЦИИ ПОЛЬЗОВАТЕЛЕЙ
# ============================================================================

# Функция для преобразования мужской фамилии в женскую
def _convert_to_female_surname(male_surname: str) -> str:
    """
    Преобразование мужской фамилии в женскую.
    
    Правила:
    - Фамилии на "ов", "ев", "ин" -> добавляем "а"
    - Фамилии на "ий", "ой" -> заменяем на "ая"
    - Остальные -> добавляем "а"
    """
    if male_surname.endswith('ий') or male_surname.endswith('ой'):
        return male_surname[:-2] + 'ая'
    elif male_surname.endswith('ов') or male_surname.endswith('ев') or male_surname.endswith('ин'):
        return male_surname + 'а'
    else:
        return male_surname + 'а'


class UserGenerator:
    """
    Класс для генерации пользователей.
    
    Генерирует пользователей с уникальными табельными номерами, ФИО и распределением по подразделениям.
    Использует библиотеки из стандартной установки Anaconda (pandas, openpyxl).
    """
    
    def __init__(
        self,
        config: Dict,
        org_data: pd.DataFrame,
        output_file_base: str,
        output_dir: str = "OUT",
        logger: Optional[logging.Logger] = None
    ) -> None:
        """
        Инициализация генератора пользователей.
        
        Args:
            config: Словарь конфигурации из LOADER_CONFIG для генератора пользователей
            org_data: DataFrame с данными организационных единиц (из листа ORG)
            output_file_base: Базовое имя выходного Excel файла (без расширения)
            output_dir: Директория для выходных файлов
            logger: Логгер для записи событий
        """
        self.config = config
        self.org_data = org_data
        self.output_file_base = output_file_base
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.logger = logger or logging.getLogger(__name__)
        
        # Настройки Excel
        self.sheet_name = config['sheet_name']
        self.max_column_width = config.get('max_column_width', 100)
        
        # Параметры генерации из конфигурации USERS
        self.business_blocks = config['business_blocks']
        self.tab_number_config = config['tab_number']
        self.gray_zone_config = config['gray_zone']
        self.fixed_distribution = config.get('fixed_distribution', {})  # Фиксированное распределение
        
        # Данные для генерации ФИО
        self.male_data = config['male_data']
        self.female_data = config['female_data']
        
        # Уникальные табельные номера
        self.used_tab_numbers: set = set()
        
        # Уникальные ФИО
        self.used_fio: set = set()
    
    def _generate_tab_number(self) -> str:
        """
        Генерация уникального табельного номера.
        
        Табельный номер: от 4 до 7 значащих цифр (больше 999), 
        записывается как текст с лидирующими нулями до 8 знаков.
        
        Returns:
            Табельный номер в формате строки с лидирующими нулями
        """
        min_val = 10 ** (self.tab_number_config['min_digits'] - 1)  # 1000
        max_val = 10 ** self.tab_number_config['max_digits'] - 1  # 9999999
        
        # Генерируем уникальный номер
        while True:
            # Генерируем случайное число от min_val до max_val
            num = random.randint(min_val, max_val)
            
            if num not in self.used_tab_numbers:
                self.used_tab_numbers.add(num)
                # Форматируем с лидирующими нулями до 8 знаков
                tab_number = str(num).zfill(self.tab_number_config['total_length'])
                return tab_number
    
    def _generate_fio(self, gender: str) -> str:
        """
        Генерация уникального ФИО.
        
        Args:
            gender: Пол ('male' или 'female')
            
        Returns:
            ФИО в формате "Фамилия Имя Отчество"
        """
        if gender == 'male':
            first_names = self.male_data['first_names']
            surnames = self.male_data['surnames']
            patronymics = self.male_data['patronymics']
        else:
            first_names = self.female_data['first_names']
            surnames = self.female_data['surnames']
            patronymics = self.female_data['patronymics']
        
        # Генерируем уникальное ФИО
        max_attempts = 10000
        for _ in range(max_attempts):
            surname = random.choice(surnames)
            first_name = random.choice(first_names)
            patronymic = random.choice(patronymics)
            
            fio = f"{surname} {first_name} {patronymic}"
            
            if fio not in self.used_fio:
                self.used_fio.add(fio)
                return fio
        
        # Если не удалось сгенерировать уникальное ФИО, добавляем номер
        surname = random.choice(surnames)
        first_name = random.choice(first_names)
        patronymic = random.choice(patronymics)
        counter = len(self.used_fio)
        fio = f"{surname} {first_name} {patronymic} ({counter})"
        self.used_fio.add(fio)
        return fio
    
    def _create_user(self, org_unit, business_block_code: str) -> Dict:
        """
        Создание одного пользователя для указанного подразделения и бизнес-блока.
        
        Args:
            org_unit: Код подразделения (может быть строкой или числом)
            business_block_code: Код бизнес-блока ('KMKKSB' или 'MNS')
            
        Returns:
            Словарь с данными пользователя
        """
        tab_number = self._generate_tab_number()
        gender = 'male' if random.random() < self.business_blocks[business_block_code]['gender_distribution'] else 'female'
        fio = self._generate_fio(gender)
        
        # Ищем подразделение, сравнивая как строки для надежности
        org_row = self.org_data[self.org_data['Код подразделения'].astype(str) == str(org_unit)].iloc[0]
        
        return {
            'Табельный номер': tab_number,
            'ФИО': fio,
            'Бизнес-блок': self.business_blocks[business_block_code]['name'],
            'Код подразделения': org_unit,
            'Код ТБ': org_row['Код ТБ'],
            'Полное ТБ': org_row['Полное ТБ'],
            'Короткое ТБ': org_row['Короткое ТБ'],
            'Код ГОСБ': org_row['Код ГОСБ'],
            'Полное ГОСБ': org_row['Полное ГОСБ'],
            'Короткое ГОСБ': org_row['Короткое ГОСБ']
        }
    
    def _distribute_users_to_org_units(self) -> List[Dict]:
        """
        Распределение пользователей по подразделениям.
        
        Правила:
        1. Сначала распределяются фиксированные количества из fixed_distribution
        2. Затем в каждом подразделении должен быть минимум 1 КМ и 1 МНС (если еще не распределено)
        3. Остальные пользователи распределяются случайно
        
        Returns:
            Список словарей с данными пользователей
        """
        users = []
        
        # Получаем список уникальных подразделений (конвертируем в строки для единообразия)
        org_units_raw = self.org_data['Код подразделения'].unique().tolist()
        # Создаем словарь для конвертации: строка -> исходное значение
        org_units_map = {str(unit): unit for unit in org_units_raw}
        org_units = list(org_units_map.keys())  # Используем строковые ключи для сравнения
        num_org_units = len(org_units)
        
        self.logger.info(f"Всего подразделений: {num_org_units}")
        self.logger.debug(f"Распределение пользователей по {num_org_units} подразделениям [class: UserGenerator | def: _distribute_users_to_org_units]")
        
        # Общее количество пользователей
        km_count = self.business_blocks['KMKKSB']['count']
        mns_count = self.business_blocks['MNS']['count']
        
        # Создаем словарь для подсчета пользователей по подразделениям (используем строковые ключи)
        org_unit_counts = {str(unit): {'KMKKSB': 0, 'MNS': 0} for unit in org_units_raw}
        
        # ШАГ 1: Распределяем фиксированные количества из fixed_distribution
        fixed_km_used = 0
        fixed_mns_used = 0
        
        if self.fixed_distribution:
            self.logger.info(f"Распределение фиксированных количеств из fixed_distribution")
            for org_unit_code, block_counts in self.fixed_distribution.items():
                # Конвертируем код в строку для сравнения
                org_unit_code_str = str(org_unit_code)
                
                # Проверяем, что подразделение существует (сравниваем как строки)
                if org_unit_code_str not in org_units:
                    self.logger.warning(f"Подразделение {org_unit_code} из fixed_distribution не найдено в данных ORG. Пропускаем.")
                    continue
                
                # Получаем исходное значение для использования в данных
                org_unit_actual = org_units_map[org_unit_code_str]
                
                # Распределяем КМ для этого подразделения
                if 'KMKKSB' in block_counts:
                    km_fixed = block_counts['KMKKSB']
                    fixed_km_used += km_fixed
                    self.logger.debug(f"Фиксированное распределение: {org_unit_code} -> {km_fixed} КМ [class: UserGenerator | def: _distribute_users_to_org_units]")
                    
                    for _ in range(km_fixed):
                        users.append(self._create_user(org_unit_actual, 'KMKKSB'))
                        org_unit_counts[org_unit_code_str]['KMKKSB'] += 1
                
                # Распределяем МНС для этого подразделения
                if 'MNS' in block_counts:
                    mns_fixed = block_counts['MNS']
                    fixed_mns_used += mns_fixed
                    self.logger.debug(f"Фиксированное распределение: {org_unit_code} -> {mns_fixed} МНС [class: UserGenerator | def: _distribute_users_to_org_units]")
                    
                    for _ in range(mns_fixed):
                        users.append(self._create_user(org_unit_actual, 'MNS'))
                        org_unit_counts[org_unit_code_str]['MNS'] += 1
            
            self.logger.info(f"Распределено фиксированных: {fixed_km_used} КМ, {fixed_mns_used} МНС")
        
        # ШАГ 2: Распределяем минимум по 1 КМ и 1 МНС в каждое подразделение (если еще не распределено)
        # Исключаем подразделения с фиксированным распределением из добавления минимума
        fixed_units_set = set()
        if self.fixed_distribution:
            for org_unit_code in self.fixed_distribution.keys():
                fixed_units_set.add(str(org_unit_code))
        
        min_km_per_unit = 1
        min_mns_per_unit = 1
        
        # Вычисляем оставшиеся количества после фиксированного распределения
        remaining_km = km_count - fixed_km_used
        remaining_mns = mns_count - fixed_mns_used
        
        # Подразделения для добавления минимума (исключаем фиксированные)
        units_for_minimum = [unit_str for unit_str in org_units if unit_str not in fixed_units_set]
        num_units_for_minimum = len(units_for_minimum)
        
        # Проверяем, достаточно ли пользователей для минимума
        min_total = (min_km_per_unit + min_mns_per_unit) * num_units_for_minimum
        if min_total > (remaining_km + remaining_mns):
            error_msg = f"Недостаточно пользователей после фиксированного распределения: требуется минимум {min_total}, доступно {remaining_km + remaining_mns}"
            self.logger.error(error_msg)
            raise ValueError(error_msg)
        
        self.logger.info(f"Распределение минимума: по {min_km_per_unit} КМ и {min_mns_per_unit} МНС в каждое подразделение (исключено {len(fixed_units_set)} с фиксированным распределением)")
        
        for org_unit_str in units_for_minimum:
            org_unit_actual = org_units_map[org_unit_str]
            
            # КМ - добавляем только если еще нет минимума
            if org_unit_counts[org_unit_str]['KMKKSB'] < min_km_per_unit:
                users.append(self._create_user(org_unit_actual, 'KMKKSB'))
                org_unit_counts[org_unit_str]['KMKKSB'] += 1
                remaining_km -= 1
            
            # МНС - добавляем только если еще нет минимума
            if org_unit_counts[org_unit_str]['MNS'] < min_mns_per_unit:
                users.append(self._create_user(org_unit_actual, 'MNS'))
                org_unit_counts[org_unit_str]['MNS'] += 1
                remaining_mns -= 1
        
        # ШАГ 3: Распределяем оставшихся пользователей случайно
        # Используем уже созданный fixed_units_set из ШАГ 2
        # Список подразделений для случайного распределения (исключаем фиксированные)
        random_org_units = [unit_str for unit_str in org_units if unit_str not in fixed_units_set]
        
        if not random_org_units:
            self.logger.warning("Все подразделения имеют фиксированное распределение. Случайное распределение невозможно.")
        else:
            self.logger.info(f"Осталось для случайного распределения: {remaining_km} КМ, {remaining_mns} МНС")
            self.logger.debug(f"Случайное распределение по {len(random_org_units)} подразделениям (исключено {len(fixed_units_set)} с фиксированным распределением) [class: UserGenerator | def: _distribute_users_to_org_units]")
            
            for _ in range(remaining_km):
                org_unit_str = random.choice(random_org_units)
                org_unit_actual = org_units_map[org_unit_str]
                users.append(self._create_user(org_unit_actual, 'KMKKSB'))
                org_unit_counts[org_unit_str]['KMKKSB'] += 1
            
            # Распределяем оставшихся МНС случайно
            for _ in range(remaining_mns):
                org_unit_str = random.choice(random_org_units)
                org_unit_actual = org_units_map[org_unit_str]
                users.append(self._create_user(org_unit_actual, 'MNS'))
                org_unit_counts[org_unit_str]['MNS'] += 1
        
        # Проверяем распределение
        km_total = sum(counts['KMKKSB'] for counts in org_unit_counts.values())
        mns_total = sum(counts['MNS'] for counts in org_unit_counts.values())
        
        self.logger.info(f"Сгенерировано пользователей: КМ={km_total}, МНС={mns_total}, всего={len(users)}")
        self.logger.debug(f"Распределение завершено. КМ: {km_total}, МНС: {mns_total} [class: UserGenerator | def: _distribute_users_to_org_units]")
        
        # Добавляем специальных пользователей "Серая зона" в каждое подразделение
        self.logger.info("Добавление специальных пользователей 'Серая зона' в каждое подразделение")
        # Используем исходные значения для специальных пользователей
        org_units_actual = [org_units_map[unit_str] for unit_str in org_units]
        users.extend(self._add_gray_zone_users(org_units_actual))
        
        self.logger.info(f"Всего пользователей после добавления 'Серая зона': {len(users)}")
        
        # Выводим статистику в DEBUG лог
        self._log_statistics(users)
        
        return users
    
    def _log_statistics(self, users: List[Dict]) -> None:
        """
        Вывод статистики по сгенерированным пользователям в DEBUG лог.
        
        Args:
            users: Список словарей с данными пользователей
        """
        df = pd.DataFrame(users)
        
        # Общая статистика
        num_blocks = df['Бизнес-блок'].nunique()
        num_tb = df['Код ТБ'].nunique()
        num_gosb = df['Код ГОСБ'].nunique()
        
        self.logger.debug(f"=== СТАТИСТИКА ПОЛЬЗОВАТЕЛЕЙ ===")
        self.logger.debug(f"Всего пользователей: {len(df)}")
        self.logger.debug(f"Количество блоков: {num_blocks}")
        self.logger.debug(f"Количество ТБ: {num_tb}")
        self.logger.debug(f"Количество ГОСБ: {num_gosb}")
        
        # Определяем пол по отчеству
        df['Пол'] = df['ФИО'].apply(lambda x: 'Мужской' if any(ending in x for ending in ['ович', 'евич', 'ич']) else 'Женский')
        
        # Статистика по блокам
        self.logger.debug(f"\n=== СТАТИСТИКА ПО БЛОКАМ ===")
        for block in df['Бизнес-блок'].unique():
            block_df = df[df['Бизнес-блок'] == block]
            male_count = len(block_df[block_df['Пол'] == 'Мужской'])
            female_count = len(block_df[block_df['Пол'] == 'Женский'])
            total = len(block_df)
            self.logger.debug(f"{block}: всего={total}, мужчин={male_count}, женщин={female_count}")
        
        # Статистика по ГОСБ (количество пользователей в каждом ГОСБ)
        self.logger.debug(f"\n=== СТАТИСТИКА ПО ГОСБ (количество пользователей) ===")
        gosb_counts = df.groupby(['Полное ГОСБ', 'Код ГОСБ']).size().sort_values(ascending=False)
        for (gosb_name, gosb_code), count in gosb_counts.items():
            self.logger.debug(f"{gosb_name} ({gosb_code}): {count} пользователей")
        
        # Статистика по блокам и ТБ
        self.logger.debug(f"\n=== СТАТИСТИКА ПО БЛОКАМ И ТБ ===")
        for (block, tb_name, tb_code), group_df in df.groupby(['Бизнес-блок', 'Полное ТБ', 'Код ТБ']):
            male_count = len(group_df[group_df['Пол'] == 'Мужской'])
            female_count = len(group_df[group_df['Пол'] == 'Женский'])
            total = len(group_df)
            self.logger.debug(f"{block} | {tb_name} ({tb_code}): всего={total}, мужчин={male_count}, женщин={female_count}")
        
        # Статистика по блокам, ТБ и ГОСБ
        self.logger.debug(f"\n=== СТАТИСТИКА ПО БЛОКАМ, ТБ И ГОСБ ===")
        for (block, tb_name, tb_code, gosb_name, gosb_code), group_df in df.groupby(['Бизнес-блок', 'Полное ТБ', 'Код ТБ', 'Полное ГОСБ', 'Код ГОСБ']):
            male_count = len(group_df[group_df['Пол'] == 'Мужской'])
            female_count = len(group_df[group_df['Пол'] == 'Женский'])
            total = len(group_df)
            self.logger.debug(f"{block} | {tb_name} ({tb_code}) | {gosb_name} ({gosb_code}): всего={total}, мужчин={male_count}, женщин={female_count}")
        
        self.logger.debug(f"=== КОНЕЦ СТАТИСТИКИ ===\n")
    
    def _add_gray_zone_users(self, org_units: List[str]) -> List[Dict]:
        """
        Добавление специальных пользователей "Серая зона" в каждое подразделение.
        
        В каждом подразделении добавляется один пользователь с:
        - Табельным номером: "gray_zone", "00000000" или "-00000001"
        - ФИО: "Серая зона" или "-"
        - Случайным распределением по блокам КМ или МНС
        
        Args:
            org_units: Список кодов подразделений
            
        Returns:
            Список словарей с данными специальных пользователей
        """
        gray_zone_users = []
        
        # Получаем параметры из централизованной конфигурации
        tab_number_options = self.gray_zone_config['tab_numbers']
        fio_options = self.gray_zone_config['fio_options']
        business_blocks_list = list(self.business_blocks.keys())
        
        for org_unit in org_units:
            # Случайно выбираем табельный номер и ФИО
            tab_number = random.choice(tab_number_options)
            fio = random.choice(fio_options)
            
            # Случайно выбираем бизнес-блок
            business_block_code = random.choice(business_blocks_list)
            business_block_name = self.business_blocks[business_block_code]['name']
            
            # Получаем данные подразделения
            org_row = self.org_data[self.org_data['Код подразделения'] == org_unit].iloc[0]
            
            gray_zone_users.append({
                'Табельный номер': tab_number,
                'ФИО': fio,
                'Бизнес-блок': business_block_name,
                'Код подразделения': org_unit,
                'Код ТБ': org_row['Код ТБ'],
                'Полное ТБ': org_row['Полное ТБ'],
                'Короткое ТБ': org_row['Короткое ТБ'],
                'Код ГОСБ': org_row['Код ГОСБ'],
                'Полное ГОСБ': org_row['Полное ГОСБ'],
                'Короткое ГОСБ': org_row['Короткое ГОСБ']
            })
        
        self.logger.info(f"Добавлено {len(gray_zone_users)} специальных пользователей 'Серая зона'")
        self.logger.debug(f"Добавлено специальных пользователей: {len(gray_zone_users)} [class: UserGenerator | def: _add_gray_zone_users]")
        
        return gray_zone_users
    
    def save_to_excel(self, users: List[Dict]) -> str:
        """
        Сохранение данных пользователей в Excel файл с настройками форматирования.
        
        Args:
            users: Список словарей с данными пользователей
            
        Returns:
            Путь к созданному файлу
        """
        # Формируем имя файла с таймштампом
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_file = f"{self.output_file_base}_{timestamp}.xlsx"
        output_path = self.output_dir / output_file
        
        self.logger.info(f"Сохранение данных пользователей в Excel: {output_file}")
        self.logger.debug(f"Сохранение {len(users)} пользователей в файл {output_file} [class: UserGenerator | def: save_to_excel]")
        
        # Создаем DataFrame
        df = pd.DataFrame(users)
        
        # Определяем порядок колонок
        column_order = [
            'Табельный номер',
            'ФИО',
            'Бизнес-блок',
            'Код подразделения',
            'Код ТБ',
            'Полное ТБ',
            'Короткое ТБ',
            'Код ГОСБ',
            'Полное ГОСБ',
            'Короткое ГОСБ'
        ]
        
        # Переупорядочиваем колонки
        df = df[column_order]
        
        # Убеждаемся, что табельный номер сохранен как текст (с лидирующими нулями)
        df['Табельный номер'] = df['Табельный номер'].astype(str)
        
        # Определяем специальных пользователей (серую зону) из централизованной конфигурации
        gray_zone_tab_numbers = self.gray_zone_config['tab_numbers']
        gray_zone_fio = self.gray_zone_config['fio_options']
        df['Специальный'] = (
            df['Табельный номер'].isin(gray_zone_tab_numbers) | 
            df['ФИО'].isin(gray_zone_fio)
        )
        
        # Разделяем на обычных и специальных пользователей
        df_normal = df[~df['Специальный']].copy()
        df_special = df[df['Специальный']].copy()
        
        # Сортируем обычных пользователей: Бизнес-блок -> Полное ТБ -> Полное ГОСБ -> ФИО
        df_normal = df_normal.sort_values(
            by=['Бизнес-блок', 'Полное ТБ', 'Полное ГОСБ', 'ФИО'],
            ascending=[True, True, True, True]
        )
        
        # Сортируем специальных пользователей: Бизнес-блок -> Полное ТБ -> Полное ГОСБ -> ФИО
        # (они будут добавлены в конец каждого ТБ)
        df_special = df_special.sort_values(
            by=['Бизнес-блок', 'Полное ТБ', 'Полное ГОСБ', 'ФИО'],
            ascending=[True, True, True, True]
        )
        
        # Объединяем: для каждого ТБ сначала обычные, потом специальные
        result_dfs = []
        for (block, tb), normal_group in df_normal.groupby(['Бизнес-блок', 'Полное ТБ']):
            # Добавляем обычных пользователей этого ТБ
            result_dfs.append(normal_group)
            
            # Добавляем специальных пользователей этого ТБ (если есть)
            special_group = df_special[
                (df_special['Бизнес-блок'] == block) & 
                (df_special['Полное ТБ'] == tb)
            ]
            if len(special_group) > 0:
                result_dfs.append(special_group)
        
        # Объединяем все группы
        df = pd.concat(result_dfs, ignore_index=True)
        
        # Удаляем временную колонку
        df = df.drop(columns=['Специальный'])
        df = df.reset_index(drop=True)
        
        self.logger.debug(f"Данные отсортированы: Бизнес-блок -> Полное ТБ -> Полное ГОСБ -> ФИО (специальные пользователи в конце каждого ТБ) [class: UserGenerator | def: save_to_excel]")
        
        # Находим последний созданный файл с таким же базовым именем
        existing_files = list(self.output_dir.glob(f"{self.output_file_base}_*.xlsx"))
        
        if existing_files:
            # Используем последний созданный файл (тот же, что и для ORG)
            excel_path = max(existing_files, key=lambda p: p.stat().st_mtime)
            self.logger.info(f"Добавление листа USERS в существующий файл: {excel_path.name}")
            file_mode = 'a'  # Добавляем лист в существующий файл
        else:
            # Создаем новый файл (не должно произойти, т.к. ORG создает файл первым)
            excel_path = self.output_dir / output_file
            self.logger.info(f"Создание нового файла: {excel_path.name}")
            file_mode = 'w'
        
        # Сохраняем в Excel
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode=file_mode, if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=self.sheet_name, index=False)
            
            # Получаем объект листа для форматирования
            worksheet = writer.sheets[self.sheet_name]
            
            # Устанавливаем формат текста для колонки "Табельный номер" (чтобы сохранить лидирующие нули)
            tab_number_col_idx = column_order.index('Табельный номер') + 1  # +1 т.к. Excel считает с 1
            tab_number_col_letter = get_column_letter(tab_number_col_idx)
            
            # Применяем текстовый формат ко всем ячейкам колонки (кроме заголовка)
            for row in range(2, len(df) + 2):  # Начинаем с 2 (пропускаем заголовок)
                cell = worksheet[f'{tab_number_col_letter}{row}']
                cell.number_format = '@'  # Текстовый формат
                # Убеждаемся, что значение записано как строка
                if isinstance(cell.value, (int, float)):
                    cell.value = str(int(cell.value)).zfill(self.tab_number_config['total_length'])
                elif isinstance(cell.value, str):
                    # Если уже строка, проверяем длину и добавляем нули если нужно
                    cell.value = cell.value.zfill(self.tab_number_config['total_length'])
            
            # Закрепляем первую строку
            worksheet.freeze_panes = 'A2'
            
            # Включаем автофильтр
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # Настраиваем ширину колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                # Находим максимальную длину содержимого в колонке
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                # Устанавливаем ширину (минимум 10, максимум max_column_width)
                adjusted_width = min(max(max_length + 2, 10), self.max_column_width)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Делаем первую строку жирной (заголовки)
            header_font = Font(bold=True)
            for cell in worksheet[1]:
                cell.font = header_font
        
        self.logger.info(f"Файл успешно создан/обновлен: {excel_path.absolute()}")
        self.logger.debug(f"Excel файл создан/обновлен: {excel_path.absolute()}, пользователей: {len(users)} [class: UserGenerator | def: save_to_excel]")
        
        return str(excel_path.absolute())
    
    def process(self) -> str:
        """
        Полный цикл генерации пользователей.
        
        Returns:
            Путь к созданному Excel файлу
        """
        self.logger.info("Начало генерации пользователей")
        self.logger.debug("Запуск полного цикла генерации пользователей [class: UserGenerator | def: process]")
        
        # Генерация пользователей
        users = self._distribute_users_to_org_units()
        
        # Сохранение
        output_file = self.save_to_excel(users)
        
        self.logger.info(f"Генерация завершена успешно. Файл: {output_file}")
        self.logger.debug(f"Генерация завершена. Создан файл: {output_file} [class: UserGenerator | def: process]")
        
        return output_file


def generate_users(
    config: Dict,
    org_data: pd.DataFrame,
    output_file_base: str = "result_base",
    output_dir: str = "OUT",
    logger: Optional[logging.Logger] = None
) -> str:
    """
    Функция для генерации пользователей.
    
    Args:
        config: Словарь конфигурации из LOADER_CONFIG для генератора пользователей
        org_data: DataFrame с данными организационных единиц (из листа ORG)
        output_file_base: Базовое имя выходного Excel файла (по умолчанию "result_base")
        output_dir: Директория для выходных файлов (по умолчанию "OUT")
        logger: Логгер для записи событий (опционально)
    
    Returns:
        Путь к созданному Excel файлу
    """
    generator = UserGenerator(
        config=config,
        org_data=org_data,
        output_file_base=output_file_base,
        output_dir=output_dir,
        logger=logger
    )
    return generator.process()


# ============================================================================
# ГЛАВНАЯ ФУНКЦИЯ
# ============================================================================

def main() -> None:
    """
    Главная функция приложения.
    
    Инициализирует логгер и выполняет основную логику программы.
    Все параметры задаются в начале файла в разделе ПАРАМЕТРЫ КОНФИГУРАЦИИ.
    """
    # Инициализация логгера
    logger = get_logger(log_dir=LOG_DIR, log_level=LOG_LEVEL)
    
    logger.info("Запуск приложения GENERATE_RANDOM_IFT_USER_12M_4Q_BASE")
    logger.debug("Инициализация основных компонентов приложения")
    
    # Сначала загружаем данные ORG (они нужны для генерации пользователей)
    org_data = None
    org_output_file = None
    
    # Загрузка и обработка данных для каждого загрузчика из конфигурации
    for loader_name, loader_config in LOADER_CONFIG.items():
        logger.info(f"Обработка загрузчика: {loader_name}")
        
        # Обработка загрузчика ORG
        if loader_name == 'ORG':
            logger.debug(f"Конфигурация загрузчика {loader_name}: входной файл={loader_config['input_file']}, лист={loader_config['sheet_name']} [class: main | def: main]")
            
            # Проверка существования входного файла
            input_path = Path(loader_config['input_file'])
            if not input_path.exists():
                error_msg = f"Входной файл не найден для загрузчика {loader_name}: {input_path.absolute()}"
                logger.error(error_msg)
                raise FileNotFoundError(error_msg)
            
            logger.info(f"Обработка файла: {input_path}")
            logger.debug(f"Входной файл: {input_path.absolute()}, базовое имя выходного: {OUTPUT_FILE_BASE}, директория выходных файлов: {OUTPUT_DIR} [class: main | def: main]")
            
            # Загрузка и обработка данных
            try:
                org_output_file = load_org_units(
                    config=loader_config,
                    output_file_base=OUTPUT_FILE_BASE,
                    output_dir=OUTPUT_DIR,
                    logger=logger
                )
                logger.info(f"Обработка загрузчика {loader_name} завершена успешно. Результат: {org_output_file}")
                
                # Загружаем данные из созданного Excel файла для использования в генераторе пользователей
                org_df = pd.read_excel(org_output_file, sheet_name=loader_config['sheet_name'])
                org_data = org_df
                logger.info(f"Загружено {len(org_df)} подразделений из листа ORG для генерации пользователей")
                
            except Exception as e:
                error_msg = f"Ошибка при обработке загрузчика {loader_name}: {str(e)}"
                logger.error(error_msg)
                raise
        
        # Обработка генератора пользователей
        elif loader_name == 'USERS':
            if org_data is None:
                error_msg = "Данные ORG не загружены. Невозможно сгенерировать пользователей."
                logger.error(error_msg)
                raise ValueError(error_msg)
            
            logger.info("Начало генерации пользователей")
            logger.debug(f"Конфигурация генератора пользователей: лист={loader_config['sheet_name']} [class: main | def: main]")
            
            try:
                # Используем тот же файл, что и для ORG
                users_output_file = generate_users(
                    config=loader_config,
                    org_data=org_data,
                    output_file_base=OUTPUT_FILE_BASE,
                    output_dir=OUTPUT_DIR,
                    logger=logger
                )
                logger.info(f"Генерация пользователей завершена успешно. Результат: {users_output_file}")
                
            except Exception as e:
                error_msg = f"Ошибка при генерации пользователей: {str(e)}"
                logger.error(error_msg)
                raise


if __name__ == "__main__":
    main()
